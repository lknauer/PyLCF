import sys as _sys, os as _os
_sys.path.insert(0, _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))))
# -*- coding: utf-8 -*-
"""Headless test: English UI builds, overlay works, sliders sensible, exports OK."""
import os, tempfile, numpy as np
import pylcf as G
import tkinter as tk
from tkinter import filedialog

# (self-contained synthetic data is built below -- no external file needed,
#  so this test runs anywhere, GitHub Actions included)
tmp = tempfile.mkdtemp()
ok = []
def check(name, cond):
    ok.append(cond); print(("PASS" if cond else "FAIL"), "-", name)

print("APP_VERSION =", G.APP_VERSION)
check("version 1.0.0", G.APP_VERSION == "1.0.0")

# --- self-contained synthetic data in the 'XY pairs' layout (no file) ---
_x = np.linspace(0.0, 100.0, 400)
def _g(c, w, a):
    return a * np.exp(-0.5 * ((_x - c) / w) ** 2)
_m1y = _g(30, 6, 1.0) + _g(70, 8, 0.6) + 0.01
_m6y = _g(45, 7, 0.8) + _g(80, 6, 0.5) + 0.01
_rng = np.random.default_rng(0)
_ny = 0.6 * _m1y + 0.4 * _m6y + 0.002 * _rng.standard_normal(_x.size)
names = ["E", "NIS", "E", "M1", "E", "M6"]
data = np.column_stack([_x, _ny, _x, _m1y, _x, _m6y])
def pair(i, j):
    x, y = data[:, i], data[:, j]; m = np.isfinite(x) & np.isfinite(y); return x[m], y[m]
nx, ny = pair(0, 1); m1x, m1y = pair(2, 3); m6x, m6y = pair(4, 5)

# --- build the English App, drive on_fit through the real GUI path ---
root = tk.Tk(); root.withdraw()
app = G.App(root)
app.spectra = [
    {"name": "NIS", "role": "measured",  "x": nx,  "y": ny,  "source": "Excel"},
    {"name": "M1",  "role": "reference", "x": m1x, "y": m1y, "source": "Excel"},
    {"name": "M6",  "role": "reference", "x": m6x, "y": m6y, "source": "Excel"},
]
app.norm.set("area"); app.mode.set("convex"); app.xmin.set("20"); app.xmax.set("")
app.on_fit()
check("on_fit set prep/primary", app.prep is not None and app.primary is not None)
check("fit R2 good (xmin=20)", app.primary.gof.r_squared > 0.95)
print("    weights:", dict(zip(app.primary.labels, np.round(app.primary.weights, 3))),
      " R2=%.3f" % app.primary.gof.r_squared)

# --- open the overlay ---
ov = G.OverlayDialog(root, app.prep, app.primary.weights, app.xlabel.get())
check("overlay has 2 sliders", len(ov.scales) == 2)
check("wmax finite & positive", all(np.isfinite(w) and w > 0 for w in ov.wmax))
_res = [w/1000.0 for w in ov.wmax]   # Tk Scale snaps to resolution=wmax/1000
check("defaults ~ auto-fit weights (within slider resolution)",
      all(abs(ov.scales[i].get() - app.primary.weights[i]) <= _res[i] + 1e-9
          for i in range(2)))
check("auto-fit weight within slider range",
      all(app.primary.weights[i] <= ov.wmax[i] + 1e-12 for i in range(2)))
print("    wmax =", [round(w, 3) for w in ov.wmax],
      " defaults =", [round(v.get(), 3) for v in ov.scales])
ov._update()
check("GoF text populated", "R-factor" in ov.gof_var.get())
check("fraction labels populated", "%" in ov.frac_labels[0].cget("text"))

# load-from-auto / reset
ov._reset()
check("reset -> all zero", all(v.get() == 0.0 for v in ov.scales))
ov._load_auto()
check("load_auto -> back to auto weights",
      np.allclose([v.get() for v in ov.scales], app.primary.weights, atol=1e-9))

# --- exports (monkeypatch the save dialog) ---
def patched(path):
    filedialog.asksaveasfilename = lambda *a, **k: path
patched(os.path.join(tmp, "ov.png")); ov._save_image()
patched(os.path.join(tmp, "ov.dat")); ov._save_data()
patched(os.path.join(tmp, "ov.csv")); ov._save_data()
patched(os.path.join(tmp, "ov.xlsx")); ov._save_xlsx()
patched(os.path.join(tmp, "ov.json")); ov._save_json()
for fn in ("ov.png", "ov.dat", "ov.csv", "ov.xlsx", "ov.json"):
    p = os.path.join(tmp, fn)
    check("export %s non-empty" % fn, os.path.exists(p) and os.path.getsize(p) > 0)
# xlsx sheets
from openpyxl import load_workbook
wb = load_workbook(os.path.join(tmp, "ov.xlsx"))
check("xlsx has fit+summary sheets", set(wb.sheetnames) == {"fit", "summary"})
# dat columns: energy, measured, fit, residual, M1, M6
_hl = [l for l in open(os.path.join(tmp, "ov.dat")).read().splitlines()
       if l.startswith("#")]
check("dat column header intact",
      any(("M1" in l and "M6" in l and "residual" in l) for l in _hl))
check("dat metadata header present",
      any(("PyLCF" in l) for l in _hl) and any(("mode:" in l) for l in _hl))

# --- regression: overlay sliders must never be clipped (bottom-packed) ---
print("\n-- overlay layout (anti-clip) regression --")
# --- regression: the slider area must always be reachable -----------------
# The plot and the controls sit in a vertical paned window with a draggable
# sash, so the user can enlarge the slider area at will and never loses access
# to the controls (this is what the cut-off-sliders bug was about).  Two panes
# in a vertical paned window means exactly one vertical sash between them, and
# that sash is draggable up/down by construction.
_pane = ov.pane
check("overlay uses a paned window", _pane.winfo_class() == "TPanedwindow")
check("paned window is vertical", str(_pane.cget("orient")) == "vertical")
check("paned window has two panes (plot + controls)", len(_pane.panes()) == 2)
_ctrl = ov.scale_w[0].master.master           # Scale -> LabelFrame -> ctrl pane
check("controls are the bottom (second) pane", str(_ctrl) == _pane.panes()[-1])
_ci = ov.canvas.get_tk_widget().pack_info()
check("plot canvas fills its pane (top+expand)",
      _ci.get("side") == "top" and str(_ci.get("expand")) in ("1", "true", "True"))
check("overlay window has a min size", ov.top.minsize()[1] >= 480)

# window height grows with the slider count, so many references still fit
import re as _re
def _ov_height(_n):
    _x = np.linspace(0, 600, 300)
    def _peak(c, w):
        return np.exp(-0.5 * ((_x - c) / w) ** 2)
    _refs = [(_x, _peak(60 + 70 * k, 18)) for k in range(_n)]
    _meas = (_x, sum(c[1] for c in _refs) / _n + 0.01)
    _p = G.prepare_arrays(_meas, _refs, ["R%d" % k for k in range(_n)], norm="area")
    _t = tk.Toplevel(root)
    _o = G.OverlayDialog(_t, _p, np.full(_n, 1.0 / _n))
    _o.top.update_idletasks()
    _h = int(_re.match(r"\d+x(\d+)", _o.top.geometry()).group(1))
    _o.top.destroy(); _t.destroy()
    return _h
_h2, _h6 = _ov_height(2), _ov_height(6)
print("    overlay height  n=2:", _h2, " n=6:", _h6)
check("overlay height grows with slider count", _h6 > _h2)

# --- slider sanity across normalizations (the user's key concern) ---
print("\n-- slider ranges per normalization --")
for nm in ("area", "max", "none"):
    prep = G.prepare_arrays((nx, ny), [(m1x, m1y), (m6x, m6y)], ["M1", "M6"],
                            norm=nm, xmin=20)
    w0 = G.run_fit(prep, "convex").weights
    o = G.OverlayDialog(root, prep, w0, "Energy")
    tot = float(np.sum(w0))
    fr = [100 * w0[i] / tot for i in range(2)]
    print("  norm=%-5s  auto w=[%s]  wmax=[%s]  default fractions=[%s]" % (
        nm,
        ", ".join("%.3g" % v for v in w0),
        ", ".join("%.3g" % v for v in o.wmax),
        ", ".join("%.1f%%" % v for v in fr)))
    check("norm=%s sliders positive & contain default" % nm,
          all(o.wmax[i] > 0 and w0[i] <= o.wmax[i] + 1e-12 for i in range(2)))
    o.top.destroy()

# --- regression: spec builders still work ---
print("\n-- import-path regression --")
# XY pairs (Excel-folder / Excel-file 'XY pairs' layout)
specs = G.specs_from_xy_pairs(names, data, ["Measured", "Reference", "Reference"],
                              ["NIS", "M1", "M6"])
check("specs_from_xy_pairs -> 3 specs, 1 measured",
      len(specs) == 3 and sum(s["role"] == "measured" for s in specs) == 1)
# paste table
txt = "E\tMeas\tRef\n0\t1\t0.5\n1\t2\t1.0\n2\t3\t1.5\n"
nm2, dat2, nsk = G.parse_table(txt)
check("parse_table 3x3", dat2.shape == (3, 3) and nsk == 0)
# shared grid
sg = G.specs_from_shared_grid(nm2, dat2, ["Energy", "Measured", "Reference"])
check("specs_from_shared_grid -> 2 specs", len(sg) == 2)

# --- copy results to clipboard ---
app._copy_results()
_clip = root.clipboard_get()
check("copy results -> clipboard non-empty", len(_clip) > 0)
check("clipboard has table + GoF",
      ("weight" in _clip) and ("R-factor" in _clip)
      and ("M1" in _clip) and ("M6" in _clip))

root.destroy()


def test_all_checks_pass():
    assert all(ok), "%d/%d checks passed" % (sum(ok), len(ok))

if __name__ == "__main__":
    raise SystemExit(0 if all(ok) else 1)
