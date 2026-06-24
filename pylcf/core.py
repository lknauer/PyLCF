#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""core.py -- PyLCF numeric core (shared by the GUI and the CLI).

Linear Combination Fitting of 1-D data: a measured y(x) is modelled as a
weighted sum of reference data sets on a common x-axis.  This module holds the
pure-numeric part (reading, resampling, normalization, the fit back-ends,
goodness-of-fit, bootstrap, F-test and the export helpers).  It imports no GUI
toolkit, so it can be used headlessly and tested without a display.

Author:  Lukas Knauer (AG Schünemann, RPTU Kaiserslautern-Landau)
License: MIT
"""
from __future__ import annotations

import warnings
from dataclasses import dataclass, asdict
from pathlib import Path

import numpy as np
from scipy.optimize import nnls, minimize

APP_VERSION = "1.0.0"
__version__ = APP_VERSION


def trapz_area(y: np.ndarray, x: np.ndarray) -> float:
    """Trapezoidal integral, compatible with numpy 1.x and 2.x."""
    try:
        return float(np.trapezoid(y, x))          # numpy >= 2.0
    except AttributeError:                          # pragma: no cover
        return float(np.trapz(y, x))                # numpy < 2.0


_COMMENT_CHARS = "#!%&*/'\""


def _sort_dedup(x, y):
    """Sort by x and drop duplicate x-values (keep first y)."""
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    x_u, idx = np.unique(x, return_index=True)
    return x_u, y[idx]


def read_spectrum(path: str, xcol: int = 0, ycol: int = 1,
                  decimal_comma: bool = False):
    """Read a 2-column spectrum (x, y) from a text file.

    Robust against header/comment lines and the common delimiters.  Lines that
    do not parse as numbers are silently skipped.  Result is sorted by x with
    duplicate x removed.
    """
    xs: list[float] = []
    ys: list[float] = []
    p = Path(path)
    if not p.is_file():
        raise FileNotFoundError(f"File not found: {path}")

    with p.open("r", encoding="utf-8-sig", errors="replace") as fh:
        for raw in fh:
            s = raw.strip()
            if not s or s[0] in _COMMENT_CHARS:
                continue
            if decimal_comma:
                # comma is the decimal separator; '.' may be a thousands sep
                s = s.replace(";", " ").replace("\t", " ")
                parts = []
                for t in s.split():
                    if "." in t and "," in t:
                        t = t.replace(".", "")
                    parts.append(t.replace(",", "."))
            else:
                s = s.replace(",", " ").replace(";", " ").replace("\t", " ")
                parts = s.split()
            if len(parts) <= max(xcol, ycol):
                continue
            try:
                x = float(parts[xcol])
                y = float(parts[ycol])
            except ValueError:
                continue
            xs.append(x)
            ys.append(y)

    if len(xs) < 2:
        raise ValueError(
            f"Could not read two numeric (x, y) rows from {Path(path).name}. "
            f"Check the column choice and decimal separator."
        )
    return _sort_dedup(xs, ys)


# --------------------------------------------------------------------------- #
#  Tabular parser for pasted-from-Excel / multi-column text
# --------------------------------------------------------------------------- #
def parse_table(text: str, delimiter: str = "auto", decimal: str = "auto"):
    """Parse a pasted/loaded data table into (col_names, data, n_skipped).

    Parameters
    ----------
    text : str
        Multi-line block.  Typically tab-separated when copied from Excel.
    delimiter : {"auto","tab","semicolon","comma","space"}
        Column separator.  "auto": tab if present, else semicolon, else
        whitespace (comma only if no whitespace columns are found).
    decimal : {"auto","point","comma"}
        Decimal separator.  "auto": comma is treated as a decimal separator
        unless the comma is itself the column delimiter.

    Returns
    -------
    names : list[str]          column names (from a header row or "Column i")
    data  : np.ndarray         shape (n_rows, n_cols), float
    n_skipped : int            number of non-empty rows that could not be parsed
    """
    text = text.lstrip("\ufeff")          # tolerate a leading UTF-8 BOM
    lines = []
    for ln in text.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        st = ln.strip()
        if not st or st[0] in _COMMENT_CHARS:
            continue
        lines.append(ln)
    if not lines:
        raise ValueError("No data rows found.")

    # ---- choose delimiter -------------------------------------------------
    if delimiter == "auto":
        if any("\t" in ln for ln in lines):
            delim = "\t"
        elif any(";" in ln for ln in lines):
            delim = ";"
        else:
            ws_cols = max(len(ln.split()) for ln in lines)
            if ws_cols >= 2:
                delim = None                     # whitespace
            elif any("," in ln for ln in lines):
                delim = ","
            else:
                delim = None
    else:
        delim = {"tab": "\t", "semicolon": ";",
                 "comma": ",", "space": None}[delimiter]

    # ---- decimal handling -------------------------------------------------
    if decimal == "auto":
        dec_comma = (delim != ",")     # comma is decimal unless it is the delimiter
    else:
        dec_comma = (decimal == "comma")

    def split_line(ln: str):
        cells = ln.split() if delim is None else ln.split(delim)
        return [c.strip() for c in cells]

    def to_float(tok: str) -> float:
        t = tok
        if dec_comma:
            if "." in t and "," in t:       # German "1.234,5" -> "1234.5"
                t = t.replace(".", "")
            t = t.replace(",", ".")
        return float(t)

    def is_number(tok: str) -> bool:
        if tok == "":
            return False
        try:
            to_float(tok)
            return True
        except ValueError:
            return False

    first = split_line(lines[0])
    has_header = any(not is_number(c) for c in first)
    if has_header:
        names = [c if c else f"Column {i + 1}" for i, c in enumerate(first)]
        data_lines = lines[1:]
        ncol = len(first)
    else:
        names = None
        data_lines = lines
        ncol = len(first)

    rows: list[list[float]] = []
    n_skipped = 0
    for ln in data_lines:
        cells = split_line(ln)
        if len(cells) < ncol:
            n_skipped += 1
            continue
        try:
            rows.append([to_float(c) for c in cells[:ncol]])
        except ValueError:
            n_skipped += 1

    if len(rows) < 2:
        raise ValueError(
            "Fewer than two valid data rows detected. Check the delimiter "
            "and decimal separator."
        )

    data = np.asarray(rows, dtype=float)
    if names is None:
        names = [f"Column {i + 1}" for i in range(ncol)]
    return names, data, n_skipped


# --------------------------------------------------------------------------- #
#  Excel (.xlsx/.xlsm) import
# --------------------------------------------------------------------------- #
def _excel_float(v) -> float:
    """Coerce an Excel cell to float. Accepts German comma strings."""
    if v is None:
        raise ValueError("empty")
    if isinstance(v, bool):                 # openpyxl may yield booleans
        raise ValueError("bool")
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        raise ValueError("empty")
    if "." in s and "," in s:               # German "1.234,5" -> "1234.5"
        s = s.replace(".", "")
    s = s.replace(",", ".")
    return float(s)


def read_excel_sheets(path):
    """Read every sheet of an .xlsx/.xlsm into {sheet_name: (names, data)}.

    Header detection mirrors :func:`parse_table` (row 0 is a header if any
    cell there is non-numeric).  Cells are coerced to float; blank or
    unparsable cells become ``NaN`` so that spectra of *different length*
    stored side by side in one sheet are handled by the spec builders.
    openpyxl is imported lazily so the numeric core has no hard dependency.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    out = {}
    try:
        for ws in wb.worksheets:
            rows = [r for r in ws.iter_rows(values_only=True)
                    if r is not None
                    and any(c is not None and str(c).strip() != "" for c in r)]
            if not rows:
                continue
            ncol = max(len(r) for r in rows)
            rows = [tuple(r) + (None,) * (ncol - len(r)) for r in rows]

            def _is_num(v):
                try:
                    _excel_float(v)
                    return True
                except (ValueError, TypeError):
                    return False

            first = rows[0]
            if any(not _is_num(v) for v in first):
                names = [str(v).strip() if (v is not None and str(v).strip())
                         else f"Column {i + 1}" for i, v in enumerate(first)]
                body = rows[1:]
            else:
                names = [f"Column {i + 1}" for i in range(ncol)]
                body = rows

            data = np.full((len(body), ncol), np.nan)
            for i, r in enumerate(body):
                for j, v in enumerate(r):
                    try:
                        data[i, j] = _excel_float(v)
                    except (ValueError, TypeError):
                        data[i, j] = np.nan
            out[ws.title] = (names, data)
    finally:
        wb.close()

    if not out:
        raise ValueError("The Excel file contains no readable data.")
    return out


def specs_from_shared_grid(names, data, roles):
    """Build specs from one Energy column + Measured/Reference columns.

    All resulting spectra share the single energy column (one common grid).
    Rows with a non-finite energy or intensity value are dropped per column.
    """
    energy_idx = [i for i, r in enumerate(roles) if r == "Energy"]
    if len(energy_idx) != 1:
        raise ValueError("Exactly one column must be marked as 'Energy'.")
    if sum(1 for r in roles if r == "Measured") > 1:
        raise ValueError("At most one column may be marked as 'Measured'.")
    used = [i for i, r in enumerate(roles) if r in ("Measured", "Reference")]
    if not used:
        raise ValueError("At least one column must be 'Measured' or 'Reference'.")

    xcol = data[:, energy_idx[0]]
    specs = []
    for i in used:
        ycol = data[:, i]
        m = np.isfinite(xcol) & np.isfinite(ycol)
        if int(m.sum()) < 2:
            raise ValueError(f"Column '{names[i]}' has too few valid values.")
        specs.append({
            "name": (names[i] or f"Column {i + 1}").strip(),
            "role": "measured" if roles[i] == "Measured" else "reference",
            "x": xcol[m].copy(), "y": ycol[m].copy(), "source": "Excel",
        })
    return specs


def specs_from_xy_pairs(names, data, pair_roles, pair_names=None):
    """Build specs from consecutive (energy, intensity) column pairs.

    Columns are grouped as (0,1), (2,3), …; each pair is one spectrum on its
    *own* grid (so the spectra may have different x/y sampling).  Rows with a
    non-finite value in a pair are dropped for that pair only, which allows
    spectra of different length to share one sheet (shorter ones padded with
    blank cells).  A trailing unpaired column is ignored.
    """
    npairs = data.shape[1] // 2
    if npairs == 0:
        raise ValueError("XY pairs need at least two columns (E, I).")
    if pair_names is None:
        pair_names = [names[2 * k + 1] if 2 * k + 1 < len(names) else f"Spectrum {k + 1}"
                      for k in range(npairs)]
    if sum(1 for r in pair_roles if r == "Measured") > 1:
        raise ValueError("At most one pair may be marked as 'Measured'.")
    used = [k for k in range(npairs) if pair_roles[k] in ("Measured", "Reference")]
    if not used:
        raise ValueError("At least one pair must be 'Measured' or 'Reference'.")

    specs = []
    for k in used:
        xcol, ycol = data[:, 2 * k], data[:, 2 * k + 1]
        m = np.isfinite(xcol) & np.isfinite(ycol)
        if int(m.sum()) < 2:
            raise ValueError(f"Pair '{pair_names[k]}' has too few valid value pairs.")
        specs.append({
            "name": (pair_names[k] or f"Spectrum {k + 1}").strip(),
            "role": "measured" if pair_roles[k] == "Measured" else "reference",
            "x": xcol[m].copy(), "y": ycol[m].copy(), "source": "Excel",
        })
    return specs


def read_xlsx_spectrum(path):
    """Read a single (energy, intensity) spectrum from an .xlsx/.xlsm file.

    Uses the first sheet and its first two columns (energy, intensity); extra
    columns are ignored.  Rows with non-finite values are dropped.  Intended
    for folders that hold one spectrum per file.  Returns (x, y).
    """
    sheets = read_excel_sheets(path)
    _names, data = next(iter(sheets.values()))      # first sheet
    if data.shape[1] < 2:
        raise ValueError("fewer than two columns (energy, intensity).")
    x, y = data[:, 0], data[:, 1]
    m = np.isfinite(x) & np.isfinite(y)
    if int(m.sum()) < 2:
        raise ValueError("too few valid (x, y) values.")
    return x[m].copy(), y[m].copy()


def specs_from_named(spectra, roles):
    """Build specs from already-read spectra plus a parallel list of roles.

    ``spectra`` is a list of dicts with ``name``/``x``/``y`` (and optional
    ``source``); ``roles`` are Measured/Reference/Ignore.  Each
    spectrum keeps its own grid (different sampling is fine).
    """
    if sum(1 for r in roles if r == "Measured") > 1:
        raise ValueError("At most one spectrum may be marked as 'Measured'.")
    used = [i for i, r in enumerate(roles) if r in ("Measured", "Reference")]
    if not used:
        raise ValueError("At least one spectrum must be 'Measured' or 'Reference'.")
    out = []
    for i in used:
        sp = spectra[i]
        out.append({
            "name": (sp["name"] or f"Spectrum {i + 1}").strip(),
            "role": "measured" if roles[i] == "Measured" else "reference",
            "x": np.asarray(sp["x"], float), "y": np.asarray(sp["y"], float),
            "source": sp.get("source", "Excel"),
        })
    return out


# --------------------------------------------------------------------------- #
#  Preparation: common grid + normalization
# --------------------------------------------------------------------------- #
@dataclass
class PreparedData:
    x: np.ndarray
    b: np.ndarray
    A: np.ndarray
    labels: list[str]
    norm: str
    raw_areas: dict[str, float]
    window: tuple[float, float]
    xname: str = "energy"


def _normalize(y: np.ndarray, x: np.ndarray, mode: str) -> np.ndarray:
    if mode == "none":
        return y
    if mode == "area":
        a = trapz_area(y, x)
        if a <= 0:
            raise ValueError(
                "Net area is <= 0, so area normalization is not meaningful "
                "(this happens when a baseline-subtracted curve has as much "
                "negative as positive area). Use 'max' or 'none' normalization, "
                "or restrict the energy window to a positive region.")
        return y / a
    if mode == "max":
        m = float(np.max(np.abs(y)))
        if m == 0:
            raise ValueError("A spectrum of all zeros cannot be max-normalized.")
        return y / m
    raise ValueError(f"Unknown normalization mode: {mode}")


def prepare_arrays(measured, subs, labels,
                   norm: str = "area",
                   xmin: float | None = None,
                   xmax: float | None = None,
                   xname: str = "energy") -> PreparedData:
    """Build a common fit grid from in-memory arrays, resample and normalize.

    ``measured`` is an (x, y) pair; ``subs`` is a list of (x, y) pairs;
    ``labels`` are their names.  Shared by the GUI and the CLI.
    """
    xname = (str(xname).strip() or "energy")
    mx, my = _sort_dedup(*measured)
    subs = [_sort_dedup(sx, sy) for sx, sy in subs]

    if len(labels) != len(subs):
        raise ValueError("Number of labels does not match number of references.")

    lo = max([mx.min()] + [sx.min() for sx, _ in subs])
    hi = min([mx.max()] + [sx.max() for sx, _ in subs])
    if xmin is not None:
        lo = max(lo, xmin)
    if xmax is not None:
        hi = min(hi, xmax)
    if not (hi > lo):
        raise ValueError(
            f"Empty fit window: overlap/limits give [{lo:.3g}, {hi:.3g}]. "
            f"Check the {xname} ranges and Emin/Emax."
        )

    mask = (mx >= lo) & (mx <= hi)
    fit_x = mx[mask]
    if fit_x.size < len(subs) + 1:
        n = max(200, 4 * fit_x.size)
        warnings.warn(
            f"Only {fit_x.size} measured point(s) fall in the fit window; "
            f"resampling onto a uniform {n}-point grid for the fit.",
            stacklevel=2)
        fit_x = np.linspace(lo, hi, n)

    b_raw = np.interp(fit_x, mx, my)
    cols = [np.interp(fit_x, sx, sy) for sx, sy in subs]

    raw_areas = {"measured": trapz_area(b_raw, fit_x)}
    for lab, c in zip(labels, cols):
        raw_areas[lab] = trapz_area(c, fit_x)

    b = _normalize(b_raw, fit_x, norm)
    A = np.column_stack([_normalize(c, fit_x, norm) for c in cols])

    return PreparedData(x=fit_x, b=b, A=A, labels=list(labels), norm=norm,
                        raw_areas=raw_areas, window=(float(lo), float(hi)),
                        xname=xname)


# --------------------------------------------------------------------------- #
#  Fitting back-ends
# --------------------------------------------------------------------------- #
def fit_linear(A, b):
    """Unconstrained least squares."""
    w, *_ = np.linalg.lstsq(A, b, rcond=None)
    return w


def fit_nnls(A, b):
    """Non-negative least squares (weights >= 0)."""
    w, _ = nnls(A, b)
    return w


def fit_convex(A, b, x0=None):
    """Non-negative weights constrained to sum to 1 (population fractions)."""
    n = A.shape[1]
    if x0 is None:
        x0 = np.full(n, 1.0 / n)
    AtA = A.T @ A
    Atb = A.T @ b

    def obj(w):
        return float(w @ (AtA @ w) - 2.0 * (Atb @ w))

    def grad(w):
        return 2.0 * (AtA @ w - Atb)

    cons = ({"type": "eq",
             "fun": lambda w: float(np.sum(w) - 1.0),
             "jac": lambda w: np.ones_like(w)},)
    bounds = [(0.0, 1.0)] * n
    res = minimize(obj, x0, jac=grad, bounds=bounds, constraints=cons,
                   method="SLSQP", options={"maxiter": 2000, "ftol": 1e-14})
    if not res.success:
        warnings.warn(f"convex fit did not fully converge: {res.message}",
                      stacklevel=2)
    w = np.clip(res.x, 0.0, None)
    s = w.sum()
    return w / s if s > 0 else w


_FITTERS = {"linear": fit_linear, "nnls": fit_nnls, "convex": fit_convex}


def _quadrature_weights(x):
    """Trapezoidal quadrature weights for a sorted 1-D grid, mean-normalized.

    Using these as per-point weights turns the unweighted sum of squares into an
    approximation of the integral int (b-fit)^2 dx, so the fit no longer depends
    on how densely each region was sampled.  Mean-normalized to 1 so the
    weighted RMSE stays comparable to the unweighted one; for a uniform grid the
    weights are ~1 everywhere.
    """
    x = np.asarray(x, float)
    n = x.size
    if n < 2:
        return np.ones(n)
    w = np.empty(n)
    w[1:-1] = (x[2:] - x[:-2]) / 2.0
    w[0] = (x[1] - x[0]) / 2.0
    w[-1] = (x[-1] - x[-2]) / 2.0
    w = np.abs(w)
    s = float(w.sum())
    return w * (n / s) if s > 0 else np.ones(n)


def _fit_weighted(mode, A, b, w=None):
    """Fit with optional per-point weights, via sqrt(w) row scaling.

    Scaling the rows of (A, b) by sqrt(w) makes ordinary lstsq / nnls / the
    convex QP minimize the *weighted* sum of squares, while the returned
    coefficients keep their original meaning.
    """
    fitter = _FITTERS[mode]
    if w is None:
        return fitter(A, b)
    s = np.sqrt(w)
    return fitter(A * s[:, None], b * s)


# --------------------------------------------------------------------------- #
#  Goodness of fit + uncertainties
# --------------------------------------------------------------------------- #
@dataclass
class GoF:
    r_factor: float
    rmse: float
    r_squared: float
    chi2_red: float | None = None


def goodness_of_fit(b, fit, sigma=None, n_params: int = 0, w=None) -> GoF:
    resid = b - fit
    if w is None:
        ss_res = float(resid @ resid)
        ss_tot = float(np.sum((b - b.mean()) ** 2))
        denom = float(b @ b)
        rmse = float(np.sqrt(ss_res / b.size))
    else:
        sw = float(np.sum(w))
        ss_res = float(np.sum(w * resid * resid))
        bw_mean = float(np.sum(w * b) / sw) if sw > 0 else 0.0
        ss_tot = float(np.sum(w * (b - bw_mean) ** 2))
        denom = float(np.sum(w * b * b))
        rmse = float(np.sqrt(ss_res / sw)) if sw > 0 else np.nan
    r_factor = ss_res / denom if denom > 0 else np.nan
    r_squared = 1.0 - ss_res / ss_tot if ss_tot > 0 else np.nan
    chi2_red = None
    if sigma is not None:
        dof = max(b.size - n_params, 1)
        chi2_red = float(np.sum((resid / sigma) ** 2) / dof)
    return GoF(r_factor=r_factor, rmse=rmse, r_squared=r_squared,
               chi2_red=chi2_red)


def bootstrap_weights(A, b, mode, n_boot, seed: int = 0, block: bool = False,
                      w=None):
    """Residual bootstrap -> (std per weight, 95% CI [low, high] per weight).

    The plain bootstrap (``block=False``) resamples individual residuals and
    assumes they are independent.  Spectral residuals are usually
    *autocorrelated* (neighbouring points are related), which makes the plain
    bootstrap underestimate the uncertainty.  ``block=True`` selects a
    moving-block bootstrap that resamples contiguous residual blocks of length
    ~sqrt(n), preserving short-range correlation and giving more honest (wider)
    intervals.
    """
    if n_boot < 2:
        raise ValueError(
            f"bootstrap needs at least 2 resamples (got {n_boot}).")
    w0 = _fit_weighted(mode, A, b, w)
    fit0 = A @ w0
    resid = b - fit0
    n = resid.size
    rng = np.random.default_rng(seed)
    draws = np.empty((n_boot, A.shape[1]))
    if block and n >= 4:
        L = max(1, int(round(np.sqrt(n))))
        nblocks = int(np.ceil(n / L))
        starts = np.arange(0, n - L + 1)
        for k in range(n_boot):
            chosen = rng.choice(starts, size=nblocks, replace=True)
            res_bb = np.concatenate([resid[s:s + L] for s in chosen])[:n]
            draws[k] = _fit_weighted(mode, A, fit0 + res_bb, w)
    else:
        for k in range(n_boot):
            draws[k] = _fit_weighted(
                mode, A, fit0 + rng.choice(resid, size=n, replace=True), w)
    std = draws.std(axis=0, ddof=1)
    ci = np.percentile(draws, [2.5, 97.5], axis=0)
    return std, ci.T


def _resid_autocorr1(resid):
    """Lag-1 autocorrelation of residuals + AR(1) effective sample size."""
    r = np.asarray(resid, float)
    n = r.size
    if n < 3:
        return 0.0, float(n)
    r = r - r.mean()
    denom = float(r @ r)
    if denom <= 0:
        return 0.0, float(n)
    rho = float((r[:-1] @ r[1:]) / denom)
    rho = max(min(rho, 0.999), -0.999)
    n_eff = n * (1.0 - rho) / (1.0 + rho)
    return rho, float(max(2.0, min(n_eff, n)))


def f_test_components(A, b, mode, rss_full=None, use_neff=False, w=None):
    """Drop-one F-test per component: does removing reference k worsen the fit?

    Compares the full model (all m references) with each reduced model that
    omits reference k (nested), returning (F, p, dof) with one entry per
    component.  df1 = 1; df2 = N - p_full, where p_full = m (linear/nnls) or
    m-1 (convex, because the sum=1 constraint removes one degree of freedom).
    A small p (< 0.05) means the component significantly improves the fit.

    IMPORTANT CAVEATS.  The F-test assumes independent, homoscedastic, Gaussian
    residuals.  Spectral residuals are autocorrelated, so the standard p-values
    are optimistic (too many components look significant); ``use_neff=True``
    substitutes an AR(1) effective sample size for N (more conservative).  The
    test is most rigorous for ``linear``; for ``nnls``/``convex`` the
    inequality/equality constraints make it approximate.
    """
    A = np.asarray(A, float)
    b = np.asarray(b, float)
    n, m = A.shape
    F = np.full(m, np.nan)
    p = np.full(m, np.nan)
    p_full = (m - 1) if mode == "convex" else m

    def wrss(cols, coef):
        r = b - A[:, cols] @ coef
        return float(np.sum(w * r * r)) if w is not None else float(r @ r)

    w_full = _fit_weighted(mode, A, b, w)
    if rss_full is None:
        rss_full = wrss(list(range(m)), w_full)
    n_use = _resid_autocorr1(b - A @ w_full)[1] if use_neff else n
    df2 = max(int(round(n_use)) - p_full, 1)
    if m < 2:
        return F, p, (1, df2)
    from scipy.stats import f as _f
    for k in range(m):
        cols = [j for j in range(m) if j != k]
        w_red = _fit_weighted(mode, A[:, cols], b, w)
        rss_red = wrss(cols, w_red)
        num = max(rss_red - rss_full, 0.0)
        den = rss_full / df2
        Fk = num / den if den > 0 else 0.0
        F[k] = Fk
        p[k] = float(_f.sf(Fk, 1, df2))
    return F, p, (1, df2)


@dataclass
class FitResult:
    mode: str
    labels: list[str]
    weights: np.ndarray
    gof: GoF
    werr: np.ndarray | None = None
    wci: np.ndarray | None = None
    fF: np.ndarray | None = None
    fp: np.ndarray | None = None
    fdof: tuple | None = None
    fp_eff: np.ndarray | None = None
    fdof_eff: tuple | None = None
    acf1: float | None = None
    n_eff: float | None = None
    weighted: bool = False

    def fit_curve(self, A):
        return A @ self.weights


def run_fit(prep: PreparedData, mode: str,
            bootstrap: int = 0, seed: int = 0,
            block: bool = False, ftest: bool = False,
            weighted: bool = False) -> FitResult:
    A, b = prep.A, prep.b
    if mode not in _FITTERS:
        raise ValueError(f"Unknown mode: {mode}")
    wq = _quadrature_weights(prep.x) if weighted else None
    coef = _fit_weighted(mode, A, b, wq)
    fit = A @ coef
    gof = goodness_of_fit(b, fit, n_params=A.shape[1], w=wq)
    werr = wci = None
    if bootstrap:
        werr, wci = bootstrap_weights(A, b, mode, bootstrap, seed,
                                      block=block, w=wq)
    fF = fp = fdof = None
    fp_eff = fdof_eff = None
    acf1 = n_eff = None
    if ftest and A.shape[1] >= 2:
        rvec = b - fit
        rss_full = (float(np.sum(wq * rvec * rvec)) if wq is not None
                    else float(rvec @ rvec))
        fF, fp, fdof = f_test_components(A, b, mode, rss_full=rss_full, w=wq)
        _, fp_eff, fdof_eff = f_test_components(
            A, b, mode, rss_full=rss_full, w=wq, use_neff=True)
        acf1, n_eff = _resid_autocorr1(rvec)
    return FitResult(mode=mode, labels=prep.labels, weights=coef,
                     gof=gof, werr=werr, wci=wci,
                     fF=fF, fp=fp, fdof=fdof, fp_eff=fp_eff, fdof_eff=fdof_eff,
                     acf1=acf1, n_eff=n_eff, weighted=weighted)


# --------------------------------------------------------------------------- #
#  Export helpers
# --------------------------------------------------------------------------- #
def _fit_columns(prep: PreparedData, res: FitResult):
    A, b, x = prep.A, prep.b, prep.x
    fit = res.fit_curve(A)
    resid = b - fit
    cols = [x, b, fit, resid] + [res.weights[i] * A[:, i]
                                 for i in range(A.shape[1])]
    names = [prep.xname, "measured", "fit", "residual"] + list(res.labels)
    return names, np.column_stack(cols)


def _fit_metadata_lines(prep, res) -> list[str]:
    """Human-readable provenance written as comment lines atop .dat/.csv."""
    total = float(res.weights.sum())
    lines = [
        f"PyLCF {APP_VERSION} - linear combination fit",
        f"mode: {res.mode}",
        f"normalization: {prep.norm}",
        f"weighting: {'dx (trapezoidal)' if res.weighted else 'uniform'}",
        f"fit window: [{prep.window[0]:.6g}, {prep.window[1]:.6g}]"
        f"  ({prep.x.size} points)",
        "weights (fraction of total):",
    ]
    for lab, w in zip(res.labels, res.weights):
        frac = 100.0 * w / total if total else float("nan")
        lines.append(f"  {lab}: {w:.6g}  ({frac:.2f} %)")
    g = res.gof
    lines.append(f"R-factor: {g.r_factor:.6e}   RMSE: {g.rmse:.6e}"
                 f"   R^2: {g.r_squared:.6f}")
    if res.werr is not None:
        lines.append("bootstrap weight std: "
                     + ", ".join(f"{lab} {e:.4g}"
                                 for lab, e in zip(res.labels, res.werr)))
    if res.fp is not None:
        def _pstr(pv):
            return "n/a" if pv is None or np.isnan(pv) else f"{pv:.3g}"
        lines.append("F-test p (full N): "
                     + ", ".join(f"{lab} {_pstr(pv)}"
                                 for lab, pv in zip(res.labels, res.fp)))
        if res.fp_eff is not None:
            lines.append("F-test p (eff. N, autocorr-adjusted): "
                         + ", ".join(f"{lab} {_pstr(pv)}"
                                     for lab, pv in zip(res.labels, res.fp_eff)))
        if res.fdof is not None:
            extra = (f"   eff. df: (1, {res.fdof_eff[1]})"
                     if res.fdof_eff is not None else "")
            lines.append(f"F-test df: (1, {res.fdof[1]}){extra}")
        if res.acf1 is not None:
            lines.append(f"residual lag-1 autocorrelation: {res.acf1:.3f}"
                         f"   effective N: {res.n_eff:.0f}")
    lines.append("")
    return lines


def export_data(prep, res, out_path: Path, delimiter: str = "\t") -> None:
    names, data = _fit_columns(prep, res)
    header = "\n".join(_fit_metadata_lines(prep, res) + [delimiter.join(names)])
    np.savetxt(out_path, data, delimiter=delimiter, header=header, comments="# ")


def _opt_float(arr, i):
    """float(arr[i]) for JSON/Excel, or None when arr is None / the value NaN."""
    if arr is None:
        return None
    v = float(arr[i])
    return None if np.isnan(v) else v


def build_json_payload(prep: PreparedData, results: list[FitResult]) -> dict:
    payload = {
        "program": f"PyLCF {APP_VERSION}",
        "normalization": prep.norm,
        "fit_window": list(prep.window),
        "n_points": int(prep.x.size),
        "raw_areas": prep.raw_areas,
        "results": [],
    }
    for r in results:
        entry = {
            "mode": r.mode,
            "weights": {lab: float(w) for lab, w in zip(r.labels, r.weights)},
            "weight_sum": float(r.weights.sum()),
            "goodness_of_fit": {k: v for k, v in asdict(r.gof).items()
                                if v is not None},
        }
        if r.werr is not None:
            entry["weight_stderr"] = {lab: float(e)
                                      for lab, e in zip(r.labels, r.werr)}
        if r.wci is not None:
            entry["weight_ci95"] = {lab: [float(r.wci[i, 0]), float(r.wci[i, 1])]
                                    for i, lab in enumerate(r.labels)}
        if r.fp is not None:
            ft = {
                "df": list(r.fdof) if r.fdof is not None else None,
                "F": {lab: _opt_float(r.fF, i)
                      for i, lab in enumerate(r.labels)},
                "p_full_N": {lab: _opt_float(r.fp, i)
                             for i, lab in enumerate(r.labels)},
            }
            if r.fp_eff is not None:
                ft["p_eff_N"] = {lab: _opt_float(r.fp_eff, i)
                                 for i, lab in enumerate(r.labels)}
                ft["df_eff"] = (list(r.fdof_eff)
                                if r.fdof_eff is not None else None)
            entry["f_test"] = ft
        if r.acf1 is not None:
            entry["residual_autocorrelation"] = {
                "lag1": float(r.acf1), "effective_N": float(r.n_eff)}
        payload["results"].append(entry)
    return payload


def export_xlsx(prep, results, primary, out_path: Path) -> None:
    """Write fit columns + a summary sheet to .xlsx (needs pandas/openpyxl)."""
    try:
        import pandas as pd
    except ImportError as exc:                                  # pragma: no cover
        raise RuntimeError(
            "The Excel export needs 'pandas' (with 'openpyxl').\n"
            "Install it with:  pip install pandas openpyxl\n"
            "Or use the .csv export instead."
        ) from exc

    names, data = _fit_columns(prep, primary)
    df = pd.DataFrame(data, columns=names)

    summ_rows = []
    for r in results:
        total = float(r.weights.sum())
        for i, lab in enumerate(r.labels):
            row = {
                "mode": r.mode,
                "component": lab,
                "weight": float(r.weights[i]),
                "percent": 100.0 * r.weights[i] / total if total else float("nan"),
                "R_factor": r.gof.r_factor,
                "R_squared": r.gof.r_squared,
                "RMSE": r.gof.rmse,
            }
            if r.werr is not None:
                row["stderr"] = float(r.werr[i])
            if r.wci is not None:
                row["ci95_low"] = float(r.wci[i, 0])
                row["ci95_high"] = float(r.wci[i, 1])
            if r.fp is not None:
                row["F_stat"] = _opt_float(r.fF, i)
                row["p_F_fullN"] = _opt_float(r.fp, i)
                if r.fp_eff is not None:
                    row["p_F_effN"] = _opt_float(r.fp_eff, i)
            if r.acf1 is not None:
                row["resid_acf1"] = float(r.acf1)
                row["effective_N"] = float(r.n_eff)
            summ_rows.append(row)
    summ = pd.DataFrame(summ_rows)

    try:
        with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
            summ.to_excel(xl, sheet_name="summary", index=False)
            df.to_excel(xl, sheet_name="fit", index=False)
    except Exception as exc:                                    # pragma: no cover
        raise RuntimeError(
            f"Could not write Excel file: {exc}\n"
            "Is 'openpyxl' installed?  pip install openpyxl"
        ) from exc




__all__ = [
    "APP_VERSION",
    "trapz_area", "read_spectrum", "parse_table", "read_excel_sheets",
    "specs_from_shared_grid", "specs_from_xy_pairs", "read_xlsx_spectrum",
    "specs_from_named", "PreparedData", "prepare_arrays",
    "fit_linear", "fit_nnls", "fit_convex", "goodness_of_fit", "GoF",
    "bootstrap_weights", "f_test_components", "FitResult", "run_fit",
    "export_data", "build_json_payload", "export_xlsx",
]
